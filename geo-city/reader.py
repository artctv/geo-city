from pathlib import Path
from typing import Generator
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from config import Config


def load_data(base_folder: Path) -> Generator[list, None, None]:
    excel_path: Path = base_folder / Config.FILES_FOLDER / Config.FILE_NAME
    workbook: Workbook = load_workbook(
        excel_path, read_only=True, data_only=True, keep_links=False
    )
    worksheet: Worksheet = workbook[Config.WORKSHEET_NAME]
    return worksheet.iter_rows(
        min_row=Config.MIN_ROW, min_col=Config.MIN_COL, max_col=Config.MAX_COL
    )


def prepare_data(line: tuple[Cell, Cell, Cell, Cell, Cell]) -> tuple[str, float, float]:
    city, _, _, lon, lat = line
    return city.value, lon.value, lat.value
