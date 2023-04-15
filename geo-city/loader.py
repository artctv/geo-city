import os
import json
from typing import Generator
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from config import Config


data_T = list[tuple[str, float, float]]


def _iterate_data() -> Generator[list, None, None]:
    workbook: Workbook = load_workbook(Config.DATA_FILE, read_only=True, data_only=True, keep_links=False)
    worksheet: Worksheet = workbook[Config.WORKSHEET_NAME]
    return worksheet.iter_rows(min_row=Config.MIN_ROW, min_col=Config.MIN_COL, max_col=Config.MAX_COL)


def _convert_data(line: tuple[Cell, Cell, Cell, Cell, Cell]) -> tuple[str, float, float]:
    city, _, _, lon, lat = line
    return city.value, lon.value, lat.value


def _load_excel_data(data_count: int = 0) -> data_T:
    data: list[tuple[str, float, float]] = []
    i: tuple[Cell, Cell, Cell, Cell, Cell]
    for n, i in enumerate(_iterate_data(), start=1):
        data.append(_convert_data(i))
        if n == data_count:
            break
    return data


def _load_temp_data(data_count: int = 0) -> data_T:
    data: list[tuple[str, float, float]] = []
    is_count_changed = False
    if Config.TEMP_FILE.is_file():
        with open(Config.TEMP_FILE, "r") as f:
            temp_data = json.load(f)
            if temp_data["data_count"] != data_count:
                is_count_changed = True
            else:
                data = temp_data["data"]

    if is_count_changed or not Config.TEMP_FILE.is_file():
        data = _load_excel_data(data_count)
        with open(Config.TEMP_FILE, "w") as f:
            temp_struct = {"data_count": data_count, "data": data}
            json.dump(temp_struct, f, ensure_ascii=False, indent=4)

    return data


def get_data(temp: bool, data_count: int) -> data_T:
    data: list[tuple[str, float, float]]
    if temp:
        os.makedirs(Config.TEMP_FOLDER, exist_ok=True)
        data = _load_temp_data(data_count)
    else:
        data = _load_excel_data(data_count)
    return data
